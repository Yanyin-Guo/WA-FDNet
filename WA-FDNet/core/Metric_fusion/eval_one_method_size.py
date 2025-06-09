import numpy as np
from PIL import Image
from Metric import *
from natsort import natsorted
from tqdm import tqdm
import os
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")


def write_excel(excel_name='metric.xlsx', worksheet_name='VIF', column_index=0, data=None):
    try:
        workbook = load_workbook(excel_name)
    except FileNotFoundError:
        workbook = Workbook()
        ws = workbook.active
        ws.title = worksheet_name

    if worksheet_name in workbook.sheetnames:
        worksheet = workbook[worksheet_name]
    else:
        worksheet = workbook.create_sheet(title=worksheet_name)

    column = get_column_letter(column_index + 1)
    for i, value in enumerate(data):
        cell = worksheet[column + str(i+1)]
        cell.value = value
        
    workbook.save(excel_name)

def evaluation_one(ir_name, vi_name, f_name):
    try:
        f_img = Image.open(f_name).convert('L')
        ir_img = Image.open(ir_name).convert('L')
        vi_img = Image.open(vi_name).convert('L')
    except:
        l = f_name.split('.')[-1]
        if l=='jpg':
            f_name = f_name.split('.')[0]+'.png'
        else:
            f_name = f_name.split('.')[0]+'.jpg'
        f_img = Image.open(f_name).convert('L')
        ir_img = Image.open(ir_name).convert('L')
        vi_img = Image.open(vi_name).convert('L')

    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_int = np.array(ir_img).astype(np.int32)
    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_int = np.array(vi_img).astype(np.int32)
    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256) ##slow

    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
    MSE = MSE_function(ir_img_double, vi_img_double, f_img_double)
    VIF = VIF_function(ir_img_double, vi_img_double, f_img_double) ##slow
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    Nabf = Nabf_function(ir_img_double, vi_img_double, f_img_double)
    SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double) ##slow
    MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double) ##very slow
    return EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM

def evaluation_one_fast(ir_name, vi_name, f_name):
    try:
        f_img = Image.open(f_name).convert('L')
        ir_img = Image.open(ir_name).convert('L')
        vi_img = Image.open(vi_name).convert('L')
    except:
        l = f_name.split('.')[-1]
        if l=='jpg':
            f_name = f_name.split('.')[0]+'.png'
        else:
            f_name = f_name.split('.')[0]+'.jpg'
        f_img = Image.open(f_name).convert('L')
        ir_img = Image.open(ir_name).convert('L')
        vi_img = Image.open(vi_name).convert('L')


    f_img_int = np.array(f_img).astype(np.int32)
    f_img_double = np.array(f_img).astype(np.float32)

    ir_img_double = np.array(ir_img).astype(np.float32)

    vi_img_double = np.array(vi_img).astype(np.float32)

    EN = EN_function(f_img_int)
    #MI = MI_function(ir_img_int, vi_img_int, f_img_int, gray_level=256) ##slow

    SF = SF_function(f_img_double)
    SD = SD_function(f_img_double)
    AG = AG_function(f_img_double)
    PSNR = PSNR_function(ir_img_double, vi_img_double, f_img_double)
    MSE = MSE_function(ir_img_double, vi_img_double, f_img_double)
    #VIF = VIF_function(ir_img_double, vi_img_double, f_img_double) ##slow
    CC = CC_function(ir_img_double, vi_img_double, f_img_double)
    SCD = SCD_function(ir_img_double, vi_img_double, f_img_double)
    Qabf = Qabf_function(ir_img_double, vi_img_double, f_img_double)
    Nabf = Nabf_function(ir_img_double, vi_img_double, f_img_double)
    #SSIM = SSIM_function(ir_img_double, vi_img_double, f_img_double) ##slow
    #MS_SSIM = MS_SSIM_function(ir_img_double, vi_img_double, f_img_double) ##very slow
    return EN, SF, AG, SD, CC, SCD, MSE, PSNR, Qabf, Nabf


# @METRIC_REGISTRY.register()
def evaluation_one_method_fast(dataset_name='NSLSR_test', data_dir='./data', result_dir='/IRFusion-main/LSFDNet/test/NSLSR/LSFDNet_320', save_dir='/IRFusion-main/LSFDNet/test/metric_LSFDNet.xlsx', Method='LSFDNet' , with_mean=True):
    EN_list = []
    SF_list = []
    AG_list = []
    SD_list = []
    CC_list = []
    SCD_list = []
    MSE_list = []
    PSNR_list = []
    Qabf_list = []
    Nabf_list = []
    filename_list = []
    ir_dir = os.path.join(data_dir, dataset_name, 'LWIR')
    vi_dir = os.path.join(data_dir, dataset_name, 'SWIR')
    # f_dir = os.path.join(result_dir, dataset_name, Method)
    f_dir = result_dir
    metric_save_name = save_dir
    filelist = natsorted(os.listdir(f_dir))
    eval_bar = tqdm(filelist)
    for _, item in enumerate(eval_bar):
        ir_name = os.path.join(ir_dir, item)
        vi_name = os.path.join(vi_dir, item)
        f_name = os.path.join(f_dir, item)
        #EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one(ir_name, vi_name, f_name)
        EN, SF, AG, SD, CC, SCD, MSE, PSNR, Qabf, Nabf = evaluation_one_fast(ir_name, vi_name, f_name)
        EN_list.append(EN)
        SF_list.append(SF)
        AG_list.append(AG)
        SD_list.append(SD)
        CC_list.append(CC)
        SCD_list.append(SCD)
        MSE_list.append(MSE)
        PSNR_list.append(PSNR)
        Qabf_list.append(Qabf)
        Nabf_list.append(Nabf)
        filename_list.append(item)
        eval_bar.set_description("{} | {}".format(Method, item))
    if with_mean:
        EN_list.append(np.mean(EN_list))
        SF_list.append(np.mean(SF_list))
        AG_list.append(np.mean(AG_list))
        SD_list.append(np.mean(SD_list))
        CC_list.append(np.mean(CC_list))
        SCD_list.append(np.mean(SCD_list))
        MSE_list.append(np.mean(MSE_list))
        PSNR_list.append(np.mean(PSNR_list))
        Qabf_list.append(np.mean(Qabf_list))
        Nabf_list.append(np.mean(Nabf_list))
        filename_list.append('mean')

        EN_list.append(np.std(EN_list))
        SF_list.append(np.std(SF_list))
        AG_list.append(np.std(AG_list))
        SD_list.append(np.std(SD_list))
        CC_list.append(np.std(CC_list[:-1]))
        SCD_list.append(np.std(SCD_list))
        MSE_list.append(np.std(MSE_list))
        PSNR_list.append(np.std(PSNR_list))
        Qabf_list.append(np.std(Qabf_list))
        Nabf_list.append(np.std(Nabf_list))
        filename_list.append('std')

    EN_list = [round(x, 3) for x in EN_list]
    SF_list = [round(x, 3) for x in SF_list]
    AG_list = [round(x, 3) for x in AG_list]
    SD_list = [round(x, 3) for x in SD_list]
    CC_list = [round(x, 3) for x in CC_list]
    SCD_list = [round(x, 3) for x in SCD_list]
    MSE_list = [round(x, 3) for x in MSE_list]
    PSNR_list = [round(x, 3) for x in PSNR_list]
    Qabf_list = [round(x, 3) for x in Qabf_list]
    Nabf_list = [round(x, 3) for x in Nabf_list]

    filename_list.insert(0, '{}'.format(Method))
    EN_list.insert(0, 'EN_list')
    SF_list.insert(0, 'SF_list')
    AG_list.insert(0, 'AG_list')
    SD_list.insert(0, 'SD_list')
    CC_list.insert(0, 'CC_list')
    SCD_list.insert(0, 'SCD_list')
    MSE_list.insert(0, 'MSE_list')
    PSNR_list.insert(0, 'PSNR_list')
    Qabf_list.insert(0, 'Qabf_list')
    Nabf_list.insert(0, 'Nabf_list')

    write_excel(metric_save_name, 'all', 0, filename_list)
    write_excel(metric_save_name, 'all', 1, EN_list)
    write_excel(metric_save_name, 'all', 2, SF_list)
    write_excel(metric_save_name, 'all', 3, AG_list)
    write_excel(metric_save_name, 'all', 4, SD_list)
    write_excel(metric_save_name, 'all', 5, CC_list)
    write_excel(metric_save_name, 'all', 6, SCD_list)
    write_excel(metric_save_name, 'all', 7, MSE_list)
    write_excel(metric_save_name, 'all', 8, PSNR_list)
    write_excel(metric_save_name, 'all', 9, Qabf_list)
    write_excel(metric_save_name, 'all', 10, Nabf_list)

    column_num=0
    write_excel(metric_save_name, 'EN', column_num, EN_list)
    write_excel(metric_save_name, 'SF', column_num, SF_list)
    write_excel(metric_save_name, 'AG', column_num, AG_list)
    write_excel(metric_save_name, 'SD', column_num, SD_list)
    write_excel(metric_save_name, 'CC', column_num, CC_list)
    write_excel(metric_save_name, 'SCD', column_num, SCD_list)
    write_excel(metric_save_name, 'MSE', column_num, MSE_list)
    write_excel(metric_save_name, 'PSNR', column_num, PSNR_list)
    write_excel(metric_save_name, 'Qabf', column_num, Qabf_list)
    write_excel(metric_save_name, 'Nabf', column_num, Nabf_list)

    return EN_list[-2:-1], SF_list[-2:-1], AG_list[-2:-1], SD_list[-2:-1], CC_list[-2:-1], SCD_list[-2:-1], MSE_list[-2:-1], PSNR_list[-2:-1], Qabf_list[-2:-1], Nabf_list[-2:-1]

def evaluation_one_method(dataset_name='NSLSR_test', data_dir='./data/', result_dir='/IRFusion-main/LSFDNet/test/NSLSR/LSFDNet_320', save_dir='/IRFusion-main/LSFDNet/test/metric_LSFDNet.xlsx', Method='LSFDNet' , with_mean=True):
    EN_list = []
    MI_list = []
    SF_list = []
    AG_list = []
    SD_list = []
    CC_list = []
    SCD_list = []
    VIF_list = []
    MSE_list = []
    PSNR_list = []
    Qabf_list = []
    Nabf_list = []
    SSIM_list = []
    MS_SSIM_list = []
    filename_list = []
    ir_dir = os.path.join(data_dir, dataset_name, 'LWIR')
    vi_dir = os.path.join(data_dir, dataset_name, 'SWIR')
    f_dir = result_dir
    metric_save_name = save_dir
    filelist = natsorted(os.listdir(f_dir))
    eval_bar = tqdm(filelist)
    for _, item in enumerate(eval_bar):
        ir_name = os.path.join(ir_dir, item)
        vi_name = os.path.join(vi_dir, item)
        f_name = os.path.join(f_dir, item)
        EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one(ir_name, vi_name, f_name)
        EN_list.append(EN)
        MI_list.append(MI)
        SF_list.append(SF)
        AG_list.append(AG)
        SD_list.append(SD)
        CC_list.append(CC)
        SCD_list.append(SCD)
        VIF_list.append(VIF)
        MSE_list.append(MSE)
        PSNR_list.append(PSNR)
        Qabf_list.append(Qabf)
        Nabf_list.append(Nabf)
        SSIM_list.append(SSIM)
        MS_SSIM_list.append(MS_SSIM)
        filename_list.append(item)
        eval_bar.set_description("{} | {}".format(Method, item))
    if with_mean:
        EN_list.append(np.mean(EN_list))
        MI_list.append(np.mean(MI_list))
        SF_list.append(np.mean(SF_list))
        AG_list.append(np.mean(AG_list))
        SD_list.append(np.mean(SD_list))
        CC_list.append(np.mean(CC_list))
        SCD_list.append(np.mean(SCD_list))
        VIF_list.append(np.mean(VIF_list))
        MSE_list.append(np.mean(MSE_list))
        PSNR_list.append(np.mean(PSNR_list))
        Qabf_list.append(np.mean(Qabf_list))
        Nabf_list.append(np.mean(Nabf_list))
        SSIM_list.append(np.mean(SSIM_list))
        MS_SSIM_list.append(np.mean(MS_SSIM_list))
        filename_list.append('mean')

        EN_list.append(np.std(EN_list))
        MI_list.append(np.std(MI_list))
        SF_list.append(np.std(SF_list))
        AG_list.append(np.std(AG_list))
        SD_list.append(np.std(SD_list))
        CC_list.append(np.std(CC_list[:-1]))
        SCD_list.append(np.std(SCD_list))
        VIF_list.append(np.std(VIF_list))
        MSE_list.append(np.std(MSE_list))
        PSNR_list.append(np.std(PSNR_list))
        Qabf_list.append(np.std(Qabf_list))
        Nabf_list.append(np.std(Nabf_list))
        SSIM_list.append(np.std(SSIM_list))
        MS_SSIM_list.append(np.std(MS_SSIM_list))
        filename_list.append('std')

    EN_list = [round(x, 3) for x in EN_list]
    MI_list = [round(x, 3) for x in MI_list]
    SF_list = [round(x, 3) for x in SF_list]
    AG_list = [round(x, 3) for x in AG_list]
    SD_list = [round(x, 3) for x in SD_list]
    CC_list = [round(x, 3) for x in CC_list]
    SCD_list = [round(x, 3) for x in SCD_list]
    VIF_list = [round(x, 3) for x in VIF_list]
    MSE_list = [round(x, 3) for x in MSE_list]
    PSNR_list = [round(x, 3) for x in PSNR_list]
    Qabf_list = [round(x, 3) for x in Qabf_list]
    Nabf_list = [round(x, 3) for x in Nabf_list]
    SSIM_list = [round(x, 3) for x in SSIM_list]
    MS_SSIM_list = [round(x, 3) for x in MS_SSIM_list]

    filename_list.insert(0, '{}'.format(Method))
    EN_list.insert(0, 'EN_list')
    MI_list.insert(0, 'MI_list')
    SF_list.insert(0, 'SF_list')
    AG_list.insert(0, 'AG_list')
    SD_list.insert(0, 'SD_list')
    CC_list.insert(0, 'CC_list')
    SCD_list.insert(0, 'SCD_list')
    VIF_list.insert(0, 'VIF_list')
    MSE_list.insert(0, 'MSE_list')
    PSNR_list.insert(0, 'PSNR_list')
    Qabf_list.insert(0, 'Qabf_list')
    Nabf_list.insert(0, 'Nabf_list')
    SSIM_list.insert(0, 'SSIM_list')
    MS_SSIM_list.insert(0, 'MS_SSIM_list')

    write_excel(metric_save_name, 'all', 0, filename_list)
    write_excel(metric_save_name, 'all', 1, EN_list)
    write_excel(metric_save_name, 'all', 2, MI_list)
    write_excel(metric_save_name, 'all', 3, SF_list)
    write_excel(metric_save_name, 'all', 4, AG_list)
    write_excel(metric_save_name, 'all', 5, SD_list)
    write_excel(metric_save_name, 'all', 6, CC_list)
    write_excel(metric_save_name, 'all', 7, SCD_list)
    write_excel(metric_save_name, 'all', 8, VIF_list)
    write_excel(metric_save_name, 'all', 9, MSE_list)
    write_excel(metric_save_name, 'all', 10, PSNR_list)
    write_excel(metric_save_name, 'all', 11, Qabf_list)
    write_excel(metric_save_name, 'all', 12, Nabf_list)
    write_excel(metric_save_name, 'all', 13, SSIM_list)
    write_excel(metric_save_name, 'all', 14, MS_SSIM_list)

    column_num=0
    write_excel(metric_save_name, 'EN', column_num, EN_list)
    write_excel(metric_save_name, 'MI', column_num, MI_list)
    write_excel(metric_save_name, 'SF', column_num, SF_list)
    write_excel(metric_save_name, 'AG', column_num, AG_list)
    write_excel(metric_save_name, 'SD', column_num, SD_list)
    write_excel(metric_save_name, 'CC', column_num, CC_list)
    write_excel(metric_save_name, 'SCD', column_num, SCD_list)
    write_excel(metric_save_name, 'VIF', column_num, VIF_list)
    write_excel(metric_save_name, 'MSE', column_num, MSE_list)
    write_excel(metric_save_name, 'PSNR', column_num, PSNR_list)
    write_excel(metric_save_name, 'Qabf', column_num, Qabf_list)
    write_excel(metric_save_name, 'Nabf', column_num, Nabf_list)
    write_excel(metric_save_name, 'SSIM', column_num, SSIM_list)
    write_excel(metric_save_name, 'MS_SSIM', column_num, MS_SSIM_list)

    return EN_list[-2:-1], MI_list[-2:-1], SF_list[-2:-1], AG_list[-2:-1], SD_list[-2:-1], CC_list[-2:-1], SCD_list[-2:-1], VIF_list[-2:-1], MSE_list[-2:-1], PSNR_list[-2:-1], Qabf_list[-2:-1], Nabf_list[-2:-1], SSIM_list[-2:-1], MS_SSIM_list[-2:-1]

if __name__ == '__main__':
    EN, MI, SF, AG, SD, CC, SCD, VIF, MSE, PSNR, Qabf, Nabf, SSIM, MS_SSIM = evaluation_one_method(result_dir='/IRFusion-main/LSFDNet/core/Metric_fusion/Results/NSLSR', save_dir='/IRFusion-main/LSFDNet/core/Metric_fusion/metric/metric_LS_320_all.xlsx', Method='LS')